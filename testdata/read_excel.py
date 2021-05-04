import sys

from openpyxl import load_workbook
from pagebase import conf
from pagebase import runlog


# 拿出excel表格中的所有数据，并且忽略表头
class ExcelTestData(object):

    def __init__(self):
        a = conf.MessageIni()
        b = a.test_data_excel()
        log = runlog.RunLog()
        try:
            self.excel = load_workbook(b[0])
        except FileNotFoundError:
            log.logError("excel文件可能不存在，请检查路径")
            sys.exit(0)
        except AttributeError:
            log.logError("excel文件可能不存在，请检查路径")
            sys.exit(0)

        try:
            self.table = self.excel[b[1]]
        except FileNotFoundError:
            log.logError("sheet页未找到，请检查名称（区分大小写）")
            sys.exit(0)
        self.rows = self.table.max_row
        self.cols = self.table.max_column

    def get_data(self):
        # 遍历所有值,包括表头
        # for row in self.table.values:
        #     for value in row:
        #         print(value)
        data = []
        for row in self.table.iter_rows(min_row=2, max_row=self.rows, min_col=2, max_col=self.cols, values_only=True):
            row = list(row)
            data.append(row)

        self.excel.close()

        return data


if __name__ == '__main__':
    a = ExcelTestData()
    print(a.get_data())
